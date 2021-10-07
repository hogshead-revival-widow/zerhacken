import shutil
from pathlib import Path
from sqlalchemy import event
from sqlalchemy.ext.hybrid import hybrid_property
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from .init import db
from .consts import ID_LENGTH, UNWEIGHTED_RESULTS, SUCCESS, NO_RESULT, MAX_TRIES
from .settings import DIR_UPLOAD
from .helper import get_uuid


class Session(db.Model):
    __tablename__ = 'session'
    id = db.Column(db.String(ID_LENGTH), primary_key=True, default=get_uuid)
    last_access = db.Column(
        db.DateTime(), server_default=db.func.now(), onupdate=db.func.now())
    files = db.relationship(
        "File", back_populates="session", cascade='all,delete')

    @classmethod
    def from_id(cls, session_id):
        return cls.query.filter_by(id=session_id).first()

    def serialize(self, **kwargs):
        files = list()
        for file in self.files:
            if 'options' in kwargs and 'exclude_files_without_results' in kwargs['options']:
                if len(file.results) == 0:
                    continue
            files.append(file.serialize(**kwargs))
        out = {
            self.id: dict(
                files=files)
        }
        return out

    def as_excel(self, **kwargs):

        exclude_files_without_results = False
        if 'options' in kwargs and 'exclude_files_without_results' in kwargs['options']:
            exclude_files_without_results = True

        wb = Workbook()
        ws_all = dict()

        for file in self.files:
            if len(file.results) > 0 and exclude_files_without_results:
                ws_all[file.id] = wb.create_sheet(file.name[:30], len(ws_all))
                ws = ws_all[file.id]
                row = len(list(ws.rows)) + 1
                ws.cell(row=row, column=1, value='Dateiname')
                ws.cell(row=row, column=2, value='Ergebnis')
                ws.cell(row=row, column=3, value='Plausibilität')
                ws.cell(row=row, column=4, value='UUID')
                row = row + 1
                ws.cell(row=row, column=1, value=file.name)
                ws.cell(row=row, column=2,
                        value=file.results[0].get_readable_plausibility())
                ws.cell(row=row, column=4, value=file.id)
                if len(file.results) > 0:
                    ws.cell(row=row, column=3,
                            value=file.results[0].plausibility)

                row = row + 1
                results_with_sequences = len(
                    [result for result in file.results if len(result.sequences) > 0])
                sequences_counter = 0
                for result in file.results:
                    if len(result.sequences) > 0:
                        sequences_counter += 1
                        row = row + 1
                        row = row + 1
                        if result.plausibility == UNWEIGHTED_RESULTS:
                            ws.cell(
                                row=row, column=1, value=f'Mögliche Menge {sequences_counter}/{results_with_sequences}')
                            row = row + 1
                        ws.cell(row=row, column=1, value='Position')
                        ws.cell(row=row, column=2, value='Start')
                        ws.cell(row=row, column=3, value='Ende')
                        ws.cell(row=row, column=4, value='Dauer')
                        for sequence in result.sequences:
                            row = row + 1
                            ws.cell(row=row, column=1, value=sequence.position)
                            ws.cell(row=row, column=2, value=sequence.start)
                            ws.cell(row=row, column=3, value=sequence.end)
                            ws.cell(row=row, column=4, value=sequence.duration)
        return save_virtual_workbook(wb)

    def get_status_all_files(self):
        status_result = list()
        status_processing = list()
        status_not_started = list()
        total_files = len(self.files)
        all_done = False

        for file in self.files:
            if len(file.results) > 0:
                status_result.append(file.id)
            if len(file.results) < 1 and not file.is_unprocessed:
                status_processing.append(file.id)
            if file.is_unprocessed:
                status_not_started.append(file.id)
        if total_files == len(status_result):
            all_done = True
        return status_result, status_processing, status_not_started, total_files, all_done


@event.listens_for(Session, 'before_delete')
def delete_files(mapper, connection, target):
    if target.files:
        shutil.rmtree(DIR_UPLOAD / target.id)


class File(db.Model):
    __tablename__ = 'file'
    id = db.Column(db.String(ID_LENGTH), primary_key=True, default=get_uuid)
    _path = db.Column(db.String(), nullable=False)
    name = db.Column(db.String(), nullable=False)
    last_access = db.Column(
        db.DateTime(), server_default=db.func.now(), onupdate=db.func.now())
    session_id = db.Column(db.String(ID_LENGTH), db.ForeignKey('session.id'))
    session = db.relationship("Session", back_populates="files", uselist=False)
    results = db.relationship(
        "Result", back_populates="file", cascade='all,delete')
    is_unprocessed = db.Column(db.Boolean(), default=True)

    def __init__(self, **kwargs):
        super(File, self).__init__(**kwargs)

    @classmethod
    def from_id(cls, file_id):
        return cls.query.filter_by(id=str(file_id)).first()

    @hybrid_property
    def path(self):
        return self._path

    @path.setter
    def path(self, value):
        if(isinstance(value, Path)):
            self._path = str(value.resolve())
        else:
            self._path = value

    def serialize(self, **kwargs):
        results = list()
        for result in self.results:
            results.append(result.serialize(**kwargs))
        out = {
            self.id: dict(
                results=results,
                file_name=self.name)
        }
        return out

    def as_excel(self):
        wb = Workbook()
        ws = wb.create_sheet(self.name[:30], 0)
        row = len(list(ws.rows)) + 1

        if len(self.results) < 0:
            ws.cell(row=row, column=1, value='Keine Ergebnisse')
            return save_virtual_workbook(wb)

        ws.cell(row=row, column=1, value='Dateiname')
        ws.cell(row=row, column=2, value='Ergebnis')
        ws.cell(row=row, column=3, value='Plausibilität')
        ws.cell(row=row, column=4, value='UUID')
        row = row + 1
        ws.cell(row=row, column=1, value=self.name)
        ws.cell(row=row, column=2,
                value=self.results[0].get_readable_plausibility())
        ws.cell(row=row, column=4, value=self.id)
        ws.cell(row=row, column=3, value=self.results[0].plausibility)
        row = row + 1
        results_with_sequences = len(
            [result for result in self.results if len(result.sequences) > 0])
        sequences_counter = 0
        for result in self.results:
            if len(result.sequences) > 0:
                sequences_counter += 1
                row = row + 1
                row = row + 1
                if result.plausibility == UNWEIGHTED_RESULTS:
                    ws.cell(
                        row=row, column=1, value=f'Mögliche Menge {sequences_counter}/{results_with_sequences}')
                    row = row + 1
                ws.cell(row=row, column=1, value='Position')
                ws.cell(row=row, column=2, value='Start')
                ws.cell(row=row, column=3, value='Ende')
                ws.cell(row=row, column=4, value='Dauer')
                for sequence in result.sequences:
                    row = row + 1
                    ws.cell(row=row, column=1, value=sequence.position)
                    ws.cell(row=row, column=2, value=sequence.start)
                    ws.cell(row=row, column=3, value=sequence.end)
                    ws.cell(row=row, column=4, value=sequence.duration)
        return save_virtual_workbook(wb)


class Result(db.Model):
    __tablename__ = 'result'
    id = db.Column(db.String(ID_LENGTH), primary_key=True, default=get_uuid)
    last_access = db.Column(
        db.DateTime(), server_default=db.func.now(), onupdate=db.func.now())
    file_id = db.Column(db.String(ID_LENGTH), db.ForeignKey('file.id'))
    file = db.relationship("File", back_populates='results', uselist=False)
    plausibility = db.Column(db.Integer, nullable=False)
    sequences = db.relationship(
        "Sequence", back_populates="result", cascade='all,delete')

    def serialize(self, **kwargs):
        sequences = list()
        for sequence in self.sequences:
            sequences.append(sequence.serialize(**kwargs))
        out = dict(
            plausibility=self.plausibility,
            sequences=sequences
        )
        return out

    def get_readable_plausibility(self):
        if self.plausibility == SUCCESS:
            return f'Zuverlässig ({self.plausibility})'
        if UNWEIGHTED_RESULTS < self.plausibility < SUCCESS:
            return f'Wahrscheinlich ({self.plausibility}/{MAX_TRIES})'
        if self.plausibility == UNWEIGHTED_RESULTS:
            return f'Vermutlich ({self.plausibility})'
        if self.plausibility == NO_RESULT:
            return 'Nichts gefunden'


class Sequence(db.Model):
    __tablename__ = 'sequence'
    id = db.Column(db.String(ID_LENGTH), primary_key=True, default=get_uuid)
    last_access = db.Column(
        db.DateTime(), server_default=db.func.now(), onupdate=db.func.now())
    result_id = db.Column(db.String(ID_LENGTH), db.ForeignKey('result.id'))
    result = db.relationship(
        "Result", back_populates='sequences', uselist=False)

    _start = db.Column(db.String(), nullable=False)
    _end = db.Column(db.String(), nullable=False)
    _duration = db.Column(db.String(), nullable=False)
    position = db.Column(db.Integer, nullable=False)

    @classmethod
    def from_dict(cls, result, start, end, duration, sequence_nr):
        sequence = Sequence(result=result, start=start,
                            end=end, duration=duration, position=sequence_nr)
        return sequence

    @hybrid_property
    def start(self):
        return self._start

    @start.setter
    def start(self, value):
        self._start = str(value)

    @hybrid_property
    def end(self):
        return self._end

    @end.setter
    def end(self, value):
        self._end = str(value)

    @hybrid_property
    def duration(self):
        return self._duration

    @duration.setter
    def duration(self, value):
        self._duration = str(value)

    def serialize(self, **kwargs):
        return dict(start=self.start, end=self.end, duration=self.duration, position=self.position)
